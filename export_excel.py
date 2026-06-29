"""
Экспорт реестра документов в Excel
"""

import os
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import registry


def export_to_excel(output_path: str = None) -> str:
    """Создаёт Excel-файл реестра документов. Возвращает путь к файлу."""

    if not output_path:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = os.path.join(os.path.dirname(__file__), f"Реестр_УЦЦП_{timestamp}.xlsx")

    docs = registry.get_all()
    stats = registry.get_stats()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Реестр документов"

    # ── Цвета ──
    COLOR_HEADER  = "1F3864"   # тёмно-синий
    COLOR_TITLE   = "2E75B6"   # синий
    COLOR_ROW_ODD = "EBF3FB"   # светло-голубой
    COLOR_WHITE   = "FFFFFF"

    # ── Заголовок отчёта ──
    ws.merge_cells("A1:G1")
    title_cell = ws["A1"]
    title_cell.value = "РЕЕСТР ДОКУМЕНТОВ — УЦЦП / Хрокой Душанбе"
    title_cell.font = Font(name="Calibri", bold=True, size=14, color=COLOR_WHITE)
    title_cell.fill = PatternFill("solid", fgColor=COLOR_HEADER)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:G2")
    sub = ws["A2"]
    sub.value = f"Дата выгрузки: {datetime.now().strftime('%d.%m.%Y %H:%M')}   |   Всего документов: {stats['total']}   |   Загружено сегодня: {stats['today']}"
    sub.font = Font(name="Calibri", italic=True, size=10, color="555555")
    sub.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 18

    # ── Шапка таблицы ──
    headers = ["№", "Номер документа", "Дата загрузки", "Кто загрузил", "Telegram", "Файл", "Ссылка"]
    col_widths = [5, 22, 20, 25, 20, 35, 15]

    for col_idx, (header, width) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.font = Font(name="Calibri", bold=True, size=11, color=COLOR_WHITE)
        cell.fill = PatternFill("solid", fgColor=COLOR_TITLE)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(
            bottom=Side(style="medium", color=COLOR_WHITE),
            right=Side(style="thin", color=COLOR_WHITE)
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[3].height = 22

    # ── Данные ──
    for row_idx, doc in enumerate(docs, start=1):
        excel_row = row_idx + 3
        fill_color = COLOR_ROW_ODD if row_idx % 2 == 0 else COLOR_WHITE

        values = [
            row_idx,
            doc["doc_number"],
            doc["uploaded_at"],
            doc["full_name"] or "—",
            f"@{doc['username']}" if doc.get("username") else "—",
            doc["filename"],
            doc["drive_url"],
        ]

        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=excel_row, column=col_idx, value=value)
            cell.font = Font(name="Calibri", size=10)
            cell.fill = PatternFill("solid", fgColor=fill_color)
            cell.alignment = Alignment(vertical="center", wrap_text=False)
            cell.border = Border(
                bottom=Side(style="thin", color="D0D0D0"),
                right=Side(style="thin", color="D0D0D0")
            )

            # Ссылка — кликабельная
            if col_idx == 7 and value and value.startswith("http"):
                cell.value = "Открыть"
                cell.hyperlink = value
                cell.font = Font(name="Calibri", size=10, color="1155CC", underline="single")

        ws.row_dimensions[excel_row].height = 18

    # ── Заморозка шапки ──
    ws.freeze_panes = "A4"

    # ── Автофильтр ──
    if docs:
        ws.auto_filter.ref = f"A3:G{len(docs) + 3}"

    wb.save(output_path)
    return output_path
